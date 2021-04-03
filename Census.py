import openpyxl
import os

os.chdir('/home/kapil/Desktop/')

wb=openpyxl.load_workbook('censuspopdata.xlsx')	# create workbook object.

active_sheet=wb.active.title # to get active worksheet in workbook.

my_dict={} # dictionary to count and store no of times census of a country take place.
count=0
li=[]  # list of all countries.

for row in range(2,wb[active_sheet].max_row+1):
	li.append(wb[active_sheet].cell(row=row,column=3).value)

# storing countries and no of times census took place.
for i in range(0,len(li)):
	count=li.count(li[i])
	my_dict[li[i]]=count
	count=0

population=0
population_dict={} # dictionary to count and store total population of each country

for key in my_dict.keys():
	for i in range(2,wb[active_sheet].max_row+1):
		
		if wb[active_sheet].cell(row=i,column=3).value==key:
			population+=wb[active_sheet].cell(row=i,column=4).value
			population_dict[key]=population
		
		else:
			population=0


file_obj=open('/home/kapil/Desktop/Population.txt','a') # to store Countries and their population in text file


for key in my_dict.keys():
	file_obj.write("Population of "+str(key)+' is: '+str(population_dict[key])+"\n\n\n")

file_obj.close()	
		
		
		

